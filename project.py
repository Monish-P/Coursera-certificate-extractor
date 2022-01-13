import fitz #pip install PyMuPDF
from openpyxl import load_workbook #pip install openpyxl
while True:
    pdffile=input("Enter the PDF file location path: ")
    doc=fitz.open(pdffile)
    page = doc.loadPage(0)
    text=page.getText()
    print(text)
    lis=text.splitlines()
    workbook=load_workbook(filename="/Users/monishpalisetti/Desktop/certificates.xlsx") #specify the path of excel sheet 
    title=''
    for c in lis[42]:
        if c not in ['*', ':', '/', '\\', '?', '[', ']']:
            title=title+c
    try:
        sheet=workbook[title]
        iter=len(sheet['A'])+1
    except Exception as e:
        sheet=workbook.create_sheet(title)
        iter=1
    student_name,course='A'+str(int(iter)),'B'+str(int(iter))
    sheet[student_name]=lis[41]
    sheet[course]=lis[42]
    workbook.save(filename="/Users/monishpalisetti/Desktop/certificates.xlsx")
    val=input("Press any key to start again or 'e' to exit: ")
    if val=='e':
        break